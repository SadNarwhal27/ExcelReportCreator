import xlrd
import openpyxl as xl
from pandas import read_excel as pd
from datetime import datetime
from pytz import timezone


# Compares 2 sets of data and finds rows that match one another
def compare_data_by_col(file):
    file += '.xlsx'
    workbook = load_workbook(file)
    sheets = load_sheets(workbook,
                         ['Data Table 1', 'Data Table 2', 'Answer 1'])

    data = []
    check = get_data(sheets[0], 1)

    for row in sheets[1].iter_rows(values_only=True):
        if row[2] in check:
            pick = int(check[row[2]])
            for temp in sheets[0].iter_rows(min_row=pick,
                                            max_row=pick,
                                            values_only=True):
                if temp[0] != row[4] or temp[4] != row[5]:
                    data.append(row)
                check.pop(row[2])

    fill_sheets(sheets[2], data)
    save_workbook(workbook, file)


# Loads an excel file when given the filename
def load_workbook(file):
    return xl.load_workbook('Raw/' + file)


# Loads up the sheets you want to work with from a workbook when given a name/names
def load_sheets(workbook, sheets):
    if type(sheets) == str:
        return workbook[sheets]
    else:
        temp = []
        for x in sheets:
            temp.append(workbook[x])
        return temp


# Fills out a given sheet with tuples from a list
def fill_sheets(sheet, data):
    for i in data:
        sheet.append(i)


# Prints out sheets in a workbook to the console
def check_sheet_names(workbook):
    print(workbook.sheetnames)


# Prints out a sheet/sheets from a workbook to the console
def check_sheet(file, sheet=None):
    test = pd(file, sheet_name=sheet)
    print(test)


# Clears out a sheet of all data
def clear_sheet(file, sheet):
    workbook = load_workbook(file)
    temp = workbook[sheet]
    temp.delete_rows(idx=1, amount=1000000)
    save_workbook(workbook, file)


# Removes a sheet from a workbook
def remove_sheet(workbook, sheet):
    workbook.remove(load_sheets(workbook, sheet))


# Creates a new sheet in a workbook
def create_sheet(workbook, sheet):
    workbook.create_sheet(sheet)


# Creates a numbered dictionary of data from a specific column in a sheet
# The data becomes the key and the row it was on becomes the value
def get_data(sheet, col):
    check = {}
    x = 1
    for row in sheet.iter_rows(values_only=True):
        check[row[col]] = str(x)
        x += 1
    return check


# Saves whatever's been done to a workbook to a new file in Reports
def save_workbook(workbook, filename):
    dt = ((timezone("US/Eastern")).localize(datetime.now())).strftime('%m%d-')
    file = 'Reports/' + dt + filename
    workbook.save(file)
    check_sheet(file)


compare_data_by_col('example')
